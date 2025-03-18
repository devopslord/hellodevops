import boto3
from openpyxl import load_workbook

# AWS Credentials (Use IAM roles or environment variables instead of hardcoding)
aws_access_key = "YOUR_ACCESS_KEY"
aws_secret_key = "YOUR_SECRET_KEY"
aws_session_token = "YOUR_SESSION_TOKEN"  # Optional for temporary credentials
aws_region = "us-east-1"  # Change to your AWS region

# Mode selection: 1 = Test Mode (No changes), 2 = Apply Mode (Makes changes)
test_mode = 1

# Initialize AWS session and EC2 client
session = boto3.Session(
    aws_access_key_id=aws_access_key,
    aws_secret_access_key=aws_secret_key,
    aws_session_token=aws_session_token,
    region_name=aws_region
)
ec2 = session.client('ec2')

# Define Excel file path and sheet name
excel_file_path = "C:/path/to/your/file.xlsx"
sheet_name = "Sheet1"  # Change if necessary

# Read volume IDs from Excel file
workbook = load_workbook(excel_file_path)
sheet = workbook[sheet_name]
volumes = [row[0] for row in sheet.iter_rows(values_only=True) if row[0]]  # Skip empty rows

# Loop through each volume
for volume_id in volumes:
    if not volume_id:
        print("[INFO] Skipping empty volume entry.")
        continue

    # Get volume details
    try:
        volume_details = ec2.describe_volumes(VolumeIds=[volume_id])['Volumes'][0]
    except Exception as e:
        print(f"[INFO] Volume {volume_id} not found or unattached: {e}")
        continue

    if volume_details.get('Attachments'):
        instance_id = volume_details['Attachments'][0]['InstanceId']

        if instance_id:
            # Get instance details
            instance = ec2.describe_instances(InstanceIds=[instance_id])['Reservations'][0]['Instances'][0]
            instance_tags = instance.get('Tags', [])

            # Check for "pcm-project_number" tag
            pcm_tag = next((tag for tag in instance_tags if tag['Key'] == 'pcm-project_number'), None)
            project_tag = next((tag for tag in instance_tags if tag['Key'] == 'project_number'), None)

            if pcm_tag:
                print(f"[INFO] Found pcm-project_number={pcm_tag['Value']} on instance {instance_id}. Applying to volume {volume_id}.")

                if test_mode == 1:
                    print(f"[TEST] Would tag volume {volume_id} with pcm-project_number={pcm_tag['Value']}")
                elif test_mode == 2:
                    ec2.create_tags(
                        Resources=[volume_id],
                        Tags=[{'Key': 'pcm-project_number', 'Value': pcm_tag['Value']}]
                    )
                    print(f"[ACTION] Tagged volume {volume_id} with pcm-project_number={pcm_tag['Value']}")
            elif project_tag:
                print(f"[INFO] Found project_number={project_tag['Value']} on instance {instance_id}. Creating pcm-project_number tag.")

                if test_mode == 1:
                    print(f"[TEST] Would create pcm-project_number={project_tag['Value']} on instance {instance_id} and volume {volume_id}")
                elif test_mode == 2:
                    # Tag the instance
                    ec2.create_tags(
                        Resources=[instance_id],
                        Tags=[{'Key': 'pcm-project_number', 'Value': project_tag['Value']}]
                    )
                    # Tag the volume
                    ec2.create_tags(
                        Resources=[volume_id],
                        Tags=[{'Key': 'pcm-project_number', 'Value': project_tag['Value']}]
                    )
                    print(f"[ACTION] Created and tagged pcm-project_number={project_tag['Value']} on instance {instance_id} and volume {volume_id}.")
            else:
                print(f"[WARNING] Instance {instance_id} has no pcm-project_number or project_number tag. Skipping.")
        else:
            print(f"[INFO] Volume {volume_id} is not attached to any instance.")
    else:
        print(f"[INFO] Volume {volume_id} is not attached to any instance.")