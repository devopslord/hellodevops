import boto3
from openpyxl import load_workbook

# AWS Credentials (Use IAM roles or environment variables instead of hardcoding)
aws_access_key = "YOUR_ACCESS_KEY"
aws_secret_key = "YOUR_SECRET_KEY"
aws_session_token = "YOUR_SESSION_TOKEN"  # Optional for temporary credentials
aws_region = "us-east-1"  # Change to your AWS region

# Mode selection: 1 = Test Mode (No changes), 2 = Apply Mode (Makes changes)
test_mode = 1

# Initialize AWS session and clients
session = boto3.Session(
    aws_access_key_id=aws_access_key,
    aws_secret_access_key=aws_secret_key,
    aws_session_token=aws_session_token,
    region_name=aws_region
)
elbv2 = session.client('elbv2')  # For Load Balancers and Target Groups
ec2 = session.client('ec2')      # For Instances and Tags

# Define Excel file path and sheet name
excel_file_path = "C:/path/to/your/loadbalancers.xlsx"
sheet_name = "Sheet1"  # Change if necessary

# Read Load Balancer ARNs from Excel file
workbook = load_workbook(excel_file_path)
sheet = workbook[sheet_name]
load_balancer_arns = [row[0] for row in sheet.iter_rows(values_only=True) if row[0]]  # Skip empty rows

# Function to get instance tags
def get_instance_tags(instance_id):
    try:
        response = ec2.describe_instances(InstanceIds=[instance_id])
        instance = response['Reservations'][0]['Instances'][0]
        return instance.get('Tags', [])
    except Exception as e:
        print(f"[ERROR] Failed to describe instance {instance_id}: {e}")
        return []

# Function to tag resources
def tag_resource(resource_arn, key, value):
    if test_mode == 1:
        print(f"[TEST] Would tag resource {resource_arn} with {key}={value}")
    elif test_mode == 2:
        try:
            elbv2.add_tags(
                ResourceArns=[resource_arn],
                Tags=[{'Key': key, 'Value': value}]
            )
            print(f"[ACTION] Tagged resource {resource_arn} with {key}={value}")
        except Exception as e:
            print(f"[ERROR] Failed to tag resource {resource_arn}: {e}")

# Loop through each Load Balancer ARN
for lb_arn in load_balancer_arns:
    print(f"[INFO] Processing Load Balancer: {lb_arn}")

    # Get target groups for the load balancer
    try:
        target_groups = elbv2.describe_target_groups(LoadBalancerArn=lb_arn)['TargetGroups']
    except Exception as e:
        print(f"[ERROR] Failed to describe target groups for Load Balancer {lb_arn}: {e}")
        continue

    for target_group in target_groups:
        target_group_arn = target_group['TargetGroupArn']
        print(f"[INFO] Processing Target Group: {target_group_arn}")

        # Get targets (instances) in the target group
        try:
            targets = elbv2.describe_target_health(TargetGroupArn=target_group_arn)['TargetHealthDescriptions']
        except Exception as e:
            print(f"[ERROR] Failed to describe target health for Target Group {target_group_arn}: {e}")
            continue

        for target in targets:
            instance_id = target['Target']['Id']
            print(f"[INFO] Processing Instance: {instance_id}")

            # Get instance tags
            instance_tags = get_instance_tags(instance_id)

            # Check for "pcm-project_number" tag
            pcm_tag = next((tag for tag in instance_tags if tag['Key'] == 'pcm-project_number'), None)
            project_tag = next((tag for tag in instance_tags if tag['Key'] == 'project_number'), None)

            if pcm_tag:
                print(f"[INFO] Found pcm-project_number={pcm_tag['Value']} on instance {instance_id}.")
                # Tag the Load Balancer if it doesn't already have the tag
                try:
                    lb_tags = elbv2.describe_tags(ResourceArns=[lb_arn])['TagDescriptions'][0]['Tags']
                    lb_pcm_tag = next((tag for tag in lb_tags if tag['Key'] == 'pcm-project_number'), None)
                    if not lb_pcm_tag:
                        tag_resource(lb_arn, 'pcm-project_number', pcm_tag['Value'])
                    else:
                        print(f"[INFO] Load Balancer {lb_arn} already has pcm-project_number={lb_pcm_tag['Value']}")
                except Exception as e:
                    print(f"[ERROR] Failed to describe tags for Load Balancer {lb_arn}: {e}")
            elif project_tag:
                print(f"[INFO] Found project_number={project_tag['Value']} on instance {instance_id}.")
                # Create "pcm-project_number" tag on the instance and Load Balancer
                tag_resource(instance_id, 'pcm-project_number', project_tag['Value'])
                tag_resource(lb_arn, 'pcm-project_number', project_tag['Value'])
            else:
                print(f"[WARNING] Instance {instance_id} has no pcm-project_number or project_number tag. Skipping.")