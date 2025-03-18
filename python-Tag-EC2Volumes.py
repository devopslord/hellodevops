import boto3
from openpyxl import load_workbook

# Set AWS credentials including session token
access_key = "YOUR_ACCESS_KEY"
secret_key = "YOUR_SECRET_KEY"
session_token = "YOUR_SESSION_TOKEN"

# Initialize a session using Amazon EC2
session = boto3.Session(
    aws_access_key_id=access_key,
    aws_secret_access_key=secret_key,
    aws_session_token=session_token,
    region_name='us-east-1'
)

# Initialize EC2 client
ec2 = session.client('ec2')

# Path to the Excel file containing the volume IDs
excel_file_path = "C:/path/to/your/volume_ids.xlsx"

# Read the volume IDs from the Excel file
workbook = load_workbook(excel_file_path)
sheet = workbook.active
volume_ids = [row[0] for row in sheet.iter_rows(values_only=True)]

# Loop through each volume ID
for volume_id in volume_ids:
    # Get the volume details
    volume = ec2.describe_volumes(VolumeIds=[volume_id])['Volumes'][0]

    # Check if the volume is attached to an instance
    if volume['Attachments']:
        instance_id = volume['Attachments'][0]['InstanceId']

        # Get the instance details
        instance = ec2.describe_instances(InstanceIds=[instance_id])['Reservations'][0]['Instances'][0]

        # Check for the "pcm-project_number" tag on the instance
        pcm_project_tag = next((tag for tag in instance.get('Tags', []) if tag['Key'] == 'pcm-project_number'), None)

        if pcm_project_tag:
            # If "pcm-project_number" tag is found, tag the volume with the same tag
            ec2.create_tags(
                Resources=[volume_id],
                Tags=[{'Key': 'pcm-project_number', 'Value': pcm_project_tag['Value']}]
            )
            print(f"Tagged volume {volume_id} with pcm-project_number: {pcm_project_tag['Value']}")
        else:
            # If "pcm-project_number" tag is not found, check for "project_number" tag
            project_tag = next((tag for tag in instance.get('Tags', []) if tag['Key'] == 'project_number'), None)

            if project_tag:
                # If "project_number" tag is found, create a new tag "pcm-project_number" on both instance and volume
                ec2.create_tags(
                    Resources=[instance_id, volume_id],
                    Tags=[{'Key': 'pcm-project_number', 'Value': project_tag['Value']}]
                )
                print(f"Tagged instance {instance_id} and volume {volume_id} with pcm-project_number: {project_tag['Value']}")
            else:
                print(f"No relevant tags found for instance {instance_id}")
    else:
        print(f"Volume {volume_id} is not attached to any instance")